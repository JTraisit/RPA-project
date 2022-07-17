
import imp
from os.path import exists
import openpyxl as xl
from datetime import datetime
import shutil


def write_excel_revenue(sys_date,sPath_Data,sPath_Template,dict_value,count_y):
    print("write revenue by lab location processing...\n")
    print("count y = :"+str(count_y))
    sys_date_split = sys_date.split("-")
    sys_day = sys_date_split[2]
    sys_month = sys_date_split[1]
    sys_year = sys_date_split[0]

    def check_file_exist(sPath_Template,sPath_Data): # use for check file Cost Outside is exist or not
        while(1):
            count = 0
            if (exists(sPath_Data+'/Output/Revenue by Lab Location_'+sys_year+'.xlsx')):
                status = 'success'
                break
            elif (not(exists(sPath_Data+'/Output/Revenue by Lab Location_'+sys_year+'.xlsx'))):
                shutil.copyfile (sPath_Template+'/Revenue by Lab Location.xlsx' , sPath_Data+'/Output/Revenue by Lab Location_'+sys_year+'.xlsx')
                count+=1
            elif(count == 2):
                status == 'fail'
                break
        return status

    def NewgetMonthMMM(date_time_str):
        datesplit = str(date_time_str).split("-")
        if ((int(datesplit[1])-1)== 0):
            date_str = (str(int(datesplit[0])-1)+'-12-'+str(datesplit[2]))
        else:
            date_str = (str(datesplit[0])+'-'+str(int(datesplit[1])-1)+'-'+str(datesplit[2]))
        date_str1 = datetime.strptime(str(date_str), '%Y-%m-%d')
        date_str2 = (date_str1.strftime('%b'))
        return date_str2

    def NewgetMonthMMM_02_31(date_time_str): 
              date_time_obj = datetime.strptime(str(date_time_str), '%Y-%m-%d')
              re_format_date = date_time_obj.strftime("%Y-%b-%d")
              re_format_date = str(re_format_date)
              date_sel = re_format_date.split("-")
              month = str(date_sel[1])
              return month

    def getLabTrue(value): # use for split value to get name of lab ex. [ALSA (L03-QC21) -> L03]
        value = value.split(" ")
        value = value[1].split("-")
        value = ((value[0])[1:])
        return value

    def replace_dict(dict):# use for coonvert all lab string to name lab. (use function getLabTrue for all lab string)
        output = {}
        keep = {}
        skip = False
        for i in dict:
            if(i == "ICAL (L10-Polyol)"):
                skip = True
                output[getLabTrue(i)] = dict.get(i)
            elif(skip):
                keep[i] = dict.get(i)
            else:
                output[getLabTrue(i)] = dict.get(i)
        for i in keep:
            if(i == "RMCL"):
                output["RSM"] = keep[i]
            else:
                output[i] = keep[i]
        return output      

    def get_lab_dir(MMMThisMonth):
        file_write =  xl.load_workbook(sPath_Data+'/Output/Revenue by Lab Location_'+sys_year+'.xlsx')
        file_write.active = file_write[MMMThisMonth]
        for sheet in file_write:
            file_write[sheet.title].views.sheetView[0].tabSelected = False
        namesheet = file_write.active
        lab_dir = {}
        y_dir = {}
        i = 4
        skip_status = False
        while(1):
            if(namesheet.cell(i,1).value == (None)):
                skip_status = True
            elif((skip_status)&(namesheet.cell(i,1).value == 'New this month/Year') ):
                y_dir[namesheet.cell(i,1).value] = i
                break
            else:
                lab_dir[namesheet.cell(i,1).value] = i
            i+=1

            
        return lab_dir,y_dir






    status = check_file_exist(sPath_Template,sPath_Data)
    if (status == 'success'):
        if((int(sys_day)>=2)&(int(sys_day)<=31)):# กรณี run วันที่ 2-31 ให้เปิดตารางเดือนนี้ (สรุปข้อมูลเดือนปัจจุบัน) 
            MMMThisMonth = NewgetMonthMMM_02_31(sys_date)
            
        elif((int(sys_day)==1)): # กรณี run วันที่ 1 ให้เปิดตารางเดือนที่แล้ว (สรุปข้อมูลเดือนที่แล้ว) 
            MMMThisMonth = NewgetMonthMMM(sys_date)
        else:
            print("Error on module write_excel_cost.py")
        labdir,Ydir = get_lab_dir(MMMThisMonth)
        labdir = replace_dict(labdir)
        try:
           file_write =  xl.load_workbook(sPath_Data+'/Output/Revenue by Lab Location_'+sys_year+'.xlsx')
        except FileNotFoundError:
            raise FileNotFoundError("File Raw Data in Master folder not found, please check your file and try again later.")
        else:
            file_write.active = file_write[MMMThisMonth]
            for sheet in file_write:
                file_write[sheet.title].views.sheetView[0].tabSelected = False
            namesheet = file_write.active
            for i in dict_value:
                namesheet.cell(labdir[i],2).value = dict_value[i]
            namesheet.cell(Ydir['New this month/Year'],2).value = count_y
            file_write.save("C:/Users/admin/Desktop/Internship/Project python version/app/data/Output/Revenue by Lab Location_"+sys_year+".xlsx")
        print("write revenue by lab location success!!!!!")
            




        