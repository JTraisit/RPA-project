#write_excel_cost
from os.path import exists
import os
import pandas as pd
import openpyxl as xl
import shutil
import numpy as np
from datetime import datetime
import xlwings as xw

def writefile(excel,excel_Name,sys_date,sPath_Data,sPath_Temp,sPath_Template):
    print("write excel cost processing...\n")
    sys_date_split = sys_date.split("-")
    sys_day = sys_date_split[2]
    sys_month = sys_date_split[1]
    sys_year = sys_date_split[0]


    def getLabTrue(value): # use for split value to get name of lab ex. [ALSA (L03-QC21) -> L03]
        value = value.split(" ")
        value = value[1].split("-")
        value = ((value[0])[1:])
        return value

    def replace_dict(dict):# use for coonvert all lab string to name lab. (use function getLabTrue for all lab string)
        output = {}
        for i in dict:
            output[getLabTrue(i)] = dict.get(i)
        return output

    def check_file_exist(sPath_Template,sPath_Data): # use for check file Cost Outside is exist or not
        while(1):
            count = 0
            if (exists(sPath_Data+'/Output/Cost Outside_'+sys_year+'.xlsx')):
                status = 'success'
                break
            elif (not(exists(sPath_Data+'/Output/Cost Outside_'+sys_year+'.xlsx'))):
                shutil.copyfile (sPath_Template+'/Cost Outside.xlsx' , sPath_Data+'/Output/Cost Outside_'+sys_year+'.xlsx')
                count+=1
            elif(count == 2):
                status == 'fail'
                break
        return status

    def get_company_dict(df): # create company in form of dictionary when key is name of company and value is number of row in excel, dictionary like this {'Apple':3,'Microsoft':2}
        # df = pd.read_excel(('C:/Users/admin/Desktop/Internship/Project python version/app/temp/Cache-2022-03-01.xlsx'),sheet_name=0)
        array = (df['u_company'].values)
        d = dict(enumerate(array.flatten(), 3))
        company_dict = {y: x for x, y in d.items()}
        return company_dict

    def NewgetMonthMMM(date_time_str):# transform a month to use ex. Run-on 1/2/2022 -> get 'Jan'
        datesplit = str(date_time_str).split("-")
        if ((int(datesplit[1])-1)== 0):
            date_str = (str(int(datesplit[0])-1)+'-12-'+str(datesplit[2])) # when Run-on 1/1/2022 -> get 'Dec'
        else:
            date_str = (str(datesplit[0])+'-'+str(int(datesplit[1])-1)+'-'+str(datesplit[2]))
        date_str1 = datetime.strptime(str(date_str), '%Y-%m-%d')
        date_str2 = (date_str1.strftime('%b'))
        return date_str2

    # def NewgetMonthMMM_02_31(date_time_str):# Convert month format to use ex. Run-on 15/2/2022 -> get 'Feb'
    #           date_time_obj = datetime.strptime(str(date_time_str), '%Y-%m-%d')
    #           re_format_date = date_time_obj.strftime("%Y-%b-%d")
    #           re_format_date = str(re_format_date)
    #           date_sel = re_format_date.split("-")
    #           month = str(date_sel[1])
    #           return month      
    def NewgetMonthMMM_02_31(date_time_str): 
              date_time_obj = datetime.strptime(str(date_time_str), '%Y-%m-%d')
              re_format_date = datetime.strptime(str(date_time_obj.date()),"%Y-%m-%d")
              month = (re_format_date.strftime('%b'))
              return month

    def NewgetMonthBeforeMMM(date_time_str):# Convert month format to use ex. Run-on 01/2/2022 -> get 'Dec' 
        datesplit = str(date_time_str).split("-")
        if ((int(datesplit[1])-1)== 0): # Run On January -> get November
            date_str = (str(int(datesplit[0])-1)+'-11-'+str(datesplit[2]))
        elif((int(datesplit[1])-2)== 0):# Run On February -> get December
            date_str = (str(int(datesplit[0])-1)+'-12-'+str(datesplit[2]))
        else:
            date_str = (str(datesplit[0])+'-'+str(int(datesplit[1])-2)+'-'+str(datesplit[2]))
        date_str1 = datetime.strptime(str(date_str), '%Y-%m-%d')
        date_str2 = (date_str1.strftime('%b'))
        return date_str2

    def get_column_lab(MMMThisMonth):# read data form Cost Outside and got header then create dictionary when key is name of the lab value is number of column like this {'L02':5,'L08':6} .
        df = pd.read_excel((sPath_Data+'/Output/Cost Outside_'+sys_year+'.xlsx'),sheet_name=MMMThisMonth)
        npa = (np.array(df.columns))# get header and convert dataframe to numpy array
        d2 = dict(enumerate(npa.flatten(), 1))
        my_dict2 = {y: x for x, y in d2.items()} #convert between key and value
        my_dict2.pop("Customer")
        status = False
        RMCL = False
        dict_keep = {}
        dict_keep2 = {}
        for i in my_dict2:
            if(not(RMCL)):
                dict_keep[i] = my_dict2.get(i)
            elif((RMCL)):
                dict_keep2[i] = my_dict2.get(i)
            if (i == 'RMCL'):
                RMCL = True
                dict_keep2["RSM"] = my_dict2.get(i)
                dict_keep.popitem()
            elif (i == 'New Customer '):# new customer with space
                status = True
            if (status):
                break
        lab_dict = replace_dict(dict_keep)
        lab_dict['RSM']=dict_keep2['RSM']
        lab_dict['Others'] = dict_keep2['Others']
        dict_keep2.pop('RSM')
        dict_keep2.pop('Others')
        return [lab_dict,dict_keep2]

    def write_name_and_quotation(df,fw,company_dict,lab_dict,MMMThisMonth): # write company name and quotation on excel cost outside
        fw.active = fw[MMMThisMonth]
        for sheet in fw:
            fw[sheet.title].views.sheetView[0].tabSelected = False
        namesheet = fw.active
        clear_contents(namesheet)# clear old contents
        # file_write.worksheets
        for index,row in df.iterrows():
           u_company,u_forlab,quotation =(row['u_company'], row['u_forlab'], row['quotation'])
           if u_forlab not in lab_dict:
              
              # print(company_dict[u_company],x["Others"],row['quotation'] )
              namesheet.cell(company_dict[u_company],lab_dict["Others"]).value = row['quotation']
           else:
              u_forlab
              # print(company_dict[u_company],x[u_forlab],row['quotation'] )
              namesheet.cell(company_dict[u_company],lab_dict[u_forlab]).value = row['quotation']
        


    def set_name_formula(fw,MMMThisMonth,company_dict,lab_dict1,lab_dict2,during_the_month_status): #set name and summary by row, column on excel
        fw.active = fw[MMMThisMonth]
        count_y = 0
        for sheet in fw:
            fw[sheet.title].views.sheetView[0].tabSelected = False
        namesheet = fw.active
        sCharacter_lab = str(chr(64+(lab_dict1["L03"])))
        sCharacter_other = str(chr(64+(lab_dict1["Others"])))
        sCharacter_Sum_Company = str(chr(64+lab_dict2['Sum Company (Customer)']))
        sCharacter_Sum_price_accumulate = str(chr(64+lab_dict2['Service Price Accumulate']))
        for i in company_dict:
            # print(my_dict[i])
            namesheet.cell(company_dict[i],1).value = (i)
            namesheet.cell(company_dict[i],lab_dict2['Sum Company (Customer)']).value = ("=SUM("+sCharacter_lab+str(company_dict[i])+":"+sCharacter_other+str(company_dict[i])+")")
            if(during_the_month_status): # when Run-on during the month
                pass
            else:
                namesheet.cell(company_dict[i],lab_dict2['Service Price Accumulate.1']).value = ("=SUM("+sCharacter_Sum_Company+str(company_dict[i])+":"+sCharacter_Sum_price_accumulate+str(company_dict[i])+")")
                if ((namesheet.cell(company_dict[i],lab_dict2['Service Price Accumulate']).value)== None):
                    namesheet.cell(company_dict[i],lab_dict2['New Customer ']).value = "Y" #write Y on column New Customer when company was firsr written.
                    count_y+=1
                    
    
        namesheet.cell(len(company_dict)+3,1).value = ("Summary")
        for i in (list(lab_dict1)):
            namesheet.cell(len(company_dict)+3,lab_dict1[i]).value = ("=SUM("+(chr(64+(lab_dict1[i])))+"3"+":"+(chr(64+(lab_dict1[i])))+str(len(company_dict)+2)+")")
        for i in (list(lab_dict2)):
            if (i !='New Customer '):
                 namesheet.cell(len(company_dict)+3,lab_dict2[i]).value = ("=SUM("+(chr(64+(lab_dict2[i])))+"3"+":"+(chr(64+(lab_dict2[i])))+str(len(company_dict)+2)+")")
            else:
                pass
        
        return count_y
               
            

    def write_sum_monthly(df,file_write2,Month,beforeMonth,lab_dict2): #write summary monthly column
        print("use write_sum_monthly ")
        file_write2.active = file_write2[Month]
        for sheet in file_write2:
            file_write2[sheet.title].views.sheetView[0].tabSelected = False
        namesheet = file_write2.active 
        # dict_comp = (get_company_dict(df))
        # print("dict_comp:"+str(len(dict_comp)))
        # print("len list(df)"+str(len(np.array(df))))
        character = str(chr(64+(lab_dict2['Service Price Accumulate.1'])))
        for i in range((len(np.array(df)))):
           namesheet.cell(i+3,15).value = "="+beforeMonth+"!"+character+str(i+3) 

    def close_file(fw):
        fw.save(sPath_Data+"/Output/Cost Outside_"+sys_year+".xlsx")

    def clear_contents(sheet): # clear content 
    # continuously delete row 3 until there
    # is only a single row left over
    # that contains column names
        while(sheet.max_row > 2):
            sheet.delete_rows(3)
        return
    def read_sum_row(company_dict,listlab,MMMThisMonth): #read summary row and create dictionary when key is name of the lab and value is result of sum a columns like this {'L02':3456.00,'L09':12345.67}
        keepsum = {}
        # listlab[1]
        wbxl = xw.Book(sPath_Data+"/Output/Cost Outside_"+sys_year+".xlsx") 
        app = xw.apps.active
        for i in (list(listlab[0])):
            keepsum[i] = wbxl.sheets[MMMThisMonth].range((chr(64+(listlab[0][i])))+str(len(company_dict)+3)).value
        wbxl.save(sPath_Data+"/Output/Cost Outside_"+sys_year+".xlsx")
        app.quit()
        

        return keepsum
    def delete_file(file_path):# delete cache file
        if  os.path.exists(file_path): #check file is existing
            os.remove(file_path)
        else:
            print("The file does not exist") # return error

    #call function
    status = check_file_exist(sPath_Template,sPath_Data) # step 1 check file exist
    if (status == 'success'): #if file cost_outside exist
        during_the_month_status = False 
        get_end_row_status = False
        print("Check file Success\n")
        if((int(sys_day)>=2)&(int(sys_day)<=31)):# กรณี run วันที่ 2-31 ให้เปิดตารางเดือนนี้ (สรุปข้อมูลเดือนปัจจุบัน) 
            MMMThisMonth = NewgetMonthMMM_02_31(sys_date)
            # during_the_month_status = True
            if((int(sys_month)>=2)&(int(sys_month)<31)):
                get_end_row_status = True #adding on 8/7/2022
            
        elif((int(sys_day)==1)): # กรณี run วันที่ 1 ให้เปิดตารางเดือนที่แล้ว (สรุปข้อมูลเดือนที่แล้ว) 
            MMMThisMonth = NewgetMonthMMM(sys_date)
            if(int(sys_month)!=2): #เนื่องจากกรณี 1/2/2022 ไม่ต้อง get row เก่า
                get_end_row_status = True
        else:
            print("Error on module write_excel_cost.py")
        dCompanany = get_company_dict(excel_Name) # create dictionary company name.
        # print(dCompanany[0])
        dLab = get_column_lab(MMMThisMonth) # create dictionary lab.
        file_write =  xl.load_workbook(sPath_Data+'/Output/Cost Outside_'+sys_year+'.xlsx') 
        data_write = write_name_and_quotation(excel,file_write,dCompanany,dLab[0],MMMThisMonth) # write company name and quotation on excel cost outside
        close_file(file_write)
        if(get_end_row_status): # when Run on day == 1 and month > 2 or month = 1 (month != 2)
            if ((len(str(int(sys_month)-1)) == 1)&(int(sys_day) == 1)):
                dataframe = pd.read_excel((sPath_Temp+'/Cache-'+sys_year+'-0'+str(int(sys_month)-1)+'-01.xlsx'),sheet_name=0)
                # delete_file((sPath_Temp+'/Cache-'+sys_year+'-0'+str(int(sys_month)-1)+'-01.xlsx'))
                print("read dataframe for add -0")
                # print(df)
            elif ((int(sys_month)-1) == 0): # #adding on 8/7/2022
                dataframe = pd.read_excel((sPath_Temp+'/Cache-'+str(int(sys_year)-1)+'-12-01.xlsx'),sheet_name=0)
                # delete_file((sPath_Temp+'/Cache-'+str(int(sys_year)-1)+'-12-01.xlsx'))
                print("read dataframe for december")
            elif((int(sys_day) > 1)&(int(sys_day) <= 31)):#adding on 8/7/2022
                dataframe = pd.read_excel((sPath_Temp+'/Cache-'+sys_year+'-'+sys_month+'-01.xlsx'),sheet_name=0)#adding on 8/7/2022
                print("read dataframe for during the month")#adding on 8/7/2022
            else:
                dataframe = pd.read_excel((sPath_Temp+'/Cache-'+sys_year+'-'+str(int(sys_month)-1)+'-01.xlsx'),sheet_name=0)
                # delete_file((sPath_Temp+'/Cache-'+sys_year+'-'+str(int(sys_month)-1)+'-01.xlsx'))
                print("read dataframe for normal")
                print(dataframe)
            file_write2 =  xl.load_workbook(sPath_Data+'/Output/Cost Outside_'+sys_year+'.xlsx')
            if(int(sys_day) == 1):#adding on 8/7/2022
                write_sum = write_sum_monthly(dataframe,file_write2,MMMThisMonth,NewgetMonthBeforeMMM(sys_date),dLab[1])
            elif((int(sys_day) > 1)&(int(sys_day) <= 31)): #adding on 8/7/2022
                write_sum = write_sum_monthly(dataframe,file_write2,MMMThisMonth,NewgetMonthMMM(sys_date),dLab[1])#adding on 8/7/2022
            else:
                print("Error date format error please check and try again")#adding on 8/7/2022
            close_file(file_write2)
            # file_write2.save("C:/Users/admin/Desktop/Internship/Project python version/app/data/Output/Cost Outside_"+sys_year+".xlsx")
        file_write1 =  xl.load_workbook(sPath_Data+'/Output/Cost Outside_'+sys_year+'.xlsx')
        count_y = set_name_formula(file_write1,MMMThisMonth,dCompanany,dLab[0],dLab[1],during_the_month_status) # ผ่าน
        close_file(file_write1)
        dsum_row = read_sum_row(dCompanany,dLab,MMMThisMonth)
    print("write excel cost success!!!!\n")
    return dsum_row,count_y